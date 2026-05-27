"""
utils/exporter.py
Generates a formatted Excel workbook of the final reconciled dataset.
Includes all computed KPI columns and any operator notes.
"""

from __future__ import annotations

import io

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_NAVY   = "1B3A5C"
_AMBER  = "C9872A"
_STRIPE = "F4F6F9"
_RED    = "FDECEA"
_GREEN  = "E8F5E9"
_THIN   = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Ordered columns to include in the export and their display labels
EXPORT_COLS: dict[str, str] = {
    "cost_center":          "Cost Center",
    "sub_dept":             "Sub-Department",
    "vendor":               "Vendor",
    "task":                 "Project Task",
    "po_number":            "PO Number",
    "pr_number":            "PR Number",
    "baseline_value":       "Baseline Value",
    "change_orders":        "Approved FCOs",
    "true_contract_value":  "True Contract Value",
    "phased_budget":        "Phased Budget",
    "sap_posted_amount":    "SAP Posted Actuals",
    "outstanding_obligation": "Outstanding Commitment",
    "total_committed":      "Total Committed",
    "remaining_budget":     "Remaining Budget",
    "utilisation_pct":      "Utilisation %",
    "notes":                "Operational Notes",
}

_MONEY_COLS = {
    "baseline_value", "change_orders", "true_contract_value", "phased_budget",
    "sap_posted_amount", "outstanding_obligation", "total_committed", "remaining_budget",
}
_AUD_FMT = '_($* #,##0.00_);[Red]_($* (#,##0.00);_($* "-"??_);_(@_)'
_PCT_FMT = "0.0%"


def export_reconciled_matrix(df: pd.DataFrame) -> bytes:
    """Return raw bytes of a formatted .xlsx workbook."""
    # Keep only columns that exist
    present = {k: v for k, v in EXPORT_COLS.items() if k in df.columns}
    export_df = df[list(present.keys())].rename(columns=present).copy()

    # Convert utilisation % from 0-100 to 0-1 for Excel percentage format
    if "Utilisation %" in export_df.columns:
        export_df["Utilisation %"] = export_df["Utilisation %"] / 100

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # Write data starting at row 3 (rows 1-2 reserved for title/blank)
        export_df.to_excel(writer, index=False, sheet_name="Reconciled Matrix", startrow=2)
        wb = writer.book
        ws = writer.sheets["Reconciled Matrix"]

        n_cols = len(present)
        last_col_letter = get_column_letter(n_cols)

        # --- Title row ---
        ws.merge_cells(f"A1:{last_col_letter}1")
        t = ws["A1"]
        t.value = "MineDept Cost & Contract Tracker  |  Reconciled Matrix Export"
        t.font = Font(bold=True, size=13, color="FFFFFF")
        t.fill = PatternFill("solid", fgColor=_NAVY)
        t.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 24

        # --- Header row (row 3, Pandas startrow=2 puts headers at row 3) ---
        money_col_indices: set[int] = set()
        pct_col_indices:   set[int] = set()
        for col_idx, (internal_key, display_name) in enumerate(present.items(), 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.font      = Font(bold=True, color="FFFFFF", size=10)
            cell.fill      = PatternFill("solid", fgColor=_AMBER)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border    = _BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(display_name) + 4)
            if internal_key in _MONEY_COLS:
                money_col_indices.add(col_idx)
            if internal_key == "utilisation_pct":
                pct_col_indices.add(col_idx)

        # --- Data rows (start at row 4) ---
        last_data_row = 3 + len(export_df)
        for row_idx in range(4, last_data_row + 1):
            row_fill = _STRIPE if row_idx % 2 == 0 else "FFFFFF"
            for col_idx in range(1, n_cols + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border    = _BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=False)
                cell.fill      = PatternFill("solid", fgColor=row_fill)
                if col_idx in money_col_indices:
                    cell.number_format = _AUD_FMT
                elif col_idx in pct_col_indices:
                    cell.number_format = _PCT_FMT
            # Highlight negative remaining-budget rows
            if "Remaining Budget" in export_df.columns:
                rb_col_idx = list(present.keys()).index("remaining_budget") + 1
                rb_cell = ws.cell(row=row_idx, column=rb_col_idx)
                try:
                    if rb_cell.value is not None and float(rb_cell.value) < 0:
                        for col_idx in range(1, n_cols + 1):
                            ws.cell(row=row_idx, column=col_idx).fill = PatternFill("solid", fgColor=_RED)
                except (TypeError, ValueError):
                    pass

        ws.freeze_panes = "A4"

    return buf.getvalue()
